"""Write screening worksheets: copy rubric template, fill step-1 score cells."""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from application_analyzer.facts import ExtractedFacts
from application_analyzer.scoring import RubricScores
from llm_score.reviewers import AgentReview

# Matches Kurup worksheet layout (see rubric/*.xlsx)
ROW_APPLICANT_NAME = 2
COL_APPLICANT_NAME = 2
ROW_TOTAL_BAND = 8
COL_SCORE_START = 7  # G — Scientific Pursuits block (cleared for step 1)
ROW_RUBRIC_ROWS = {
    "scientific_pursuits_education": 5,
    "scientific_pursuits_output": 6,
    "professional_leadership_education": 7,
    "professional_leadership_output": 8,
    "social_leadership": 9,
    "resilience": 10,
    "endorsement": 11,
    "reviewer_recommendation": 12,
    "medical_school_quality": 14,
    "medical_school_performance": 15,
    "undergraduate_quality": 16,
    "undergraduate_performance": 17,
    "usmle_step1": 18,
}
COL_DOC_A = 4  # D
COL_DOC_B = 5  # E
ROW_SCORE_HEADERS = 4
COL_SUMMARY = {"msq": 14, "msp": 15, "uq": 16, "up": 17, "usmle": 18}
PURPLE_FILL = PatternFill(fill_type="solid", fgColor="D9B3FF")
AGENT_FILL = PatternFill(fill_type="solid", fgColor="B3D9FF")

STEP1_ROWS = frozenset(
    {
        "medical_school_quality",
        "medical_school_performance",
        "undergraduate_quality",
        "undergraduate_performance",
        "usmle_step1",
    }
)
SUBJECTIVE_ROWS = frozenset(ROW_RUBRIC_ROWS) - STEP1_ROWS

# Matches original D2 total: subjective rows + objective rows; excludes rec (12) and USMLE (18).
_TOTAL_FORMULA_ROWS = (5, 6, 7, 8, 9, 10, 11, 14, 15, 16, 17)


def _safe_sheet_title(name: str) -> str:
    bad = r'[:\\/?*\[\]]'
    t = re.sub(bad, "-", name).strip()
    return (t[:31] if len(t) > 31 else t) or "Applicant"


def _copy_sheet(wb: Workbook, source: Worksheet, title: str) -> Worksheet:
    """Duplicate a worksheet preserving cell styles (openpyxl copy_worksheet)."""
    target = wb.copy_worksheet(source)
    target.title = _safe_sheet_title(title)
    return target


def _total_formula(column_letter: str) -> str:
    terms = "+".join(f"{column_letter}{r}" for r in _TOTAL_FORMULA_ROWS)
    return f"={terms}"


def _setup_reviewer_columns(ws: Worksheet) -> None:
    ws.cell(row=ROW_SCORE_HEADERS, column=COL_DOC_A).value = "Doc A"
    ws.cell(row=ROW_SCORE_HEADERS, column=COL_DOC_B).value = "Doc B"
    ws.cell(row=ROW_APPLICANT_NAME, column=COL_DOC_A).value = _total_formula("D")
    ws.cell(row=ROW_APPLICANT_NAME, column=COL_DOC_B).value = _total_formula("E")


def _clear_reviewer_scores(ws: Worksheet) -> None:
    """Clear subjective rubric rows; step 1 fills objective rows later."""
    for r in range(5, 13):
        ws.cell(row=r, column=COL_DOC_A).value = None
        ws.cell(row=r, column=COL_DOC_B).value = None
    ws.cell(row=ROW_TOTAL_BAND, column=6).value = None  # F8 total
    for c in range(COL_SCORE_START, COL_SUMMARY["msq"]):
        ws.cell(row=ROW_TOTAL_BAND, column=c).value = None


def _set_automated_score(
    ws: Worksheet,
    row: int,
    value: int | float | str | None,
) -> None:
    """Pre-fill both reviewer columns with the same rule-based step-1 score."""
    for col in (COL_DOC_A, COL_DOC_B):
        cell = ws.cell(row=row, column=col)
        cell.value = value
        if value is not None:
            cell.fill = PURPLE_FILL


def _set_agent_score(
    ws: Worksheet,
    row: int,
    column: int,
    value: int | float | str | None,
) -> None:
    cell = ws.cell(row=row, column=column)
    cell.value = value
    if value is not None:
        cell.fill = AGENT_FILL


def _fill_agent_review(ws: Worksheet, column: int, review: AgentReview | None) -> None:
    if review is None:
        return
    for key in SUBJECTIVE_ROWS:
        row = ROW_RUBRIC_ROWS[key]
        _set_agent_score(ws, row, column, review.scores.get(key))
    if review.summary:
        existing = ws.cell(row=13, column=COL_APPLICANT_NAME).value
        prefix = f"{review.agent_name}: {review.summary}"
        ws.cell(row=13, column=COL_APPLICANT_NAME).value = (
            f"{existing}\n\n{prefix}" if existing else prefix
        )


def _set_summary_cell(
    ws: Worksheet,
    column: int,
    value: int | float | str | None,
) -> None:
    cell = ws.cell(row=ROW_TOTAL_BAND, column=column)
    cell.value = value
    if value is not None:
        cell.fill = PURPLE_FILL


def _strip_original_sheets(wb: Workbook, original_names: list[str]) -> None:
    for name in original_names:
        if name in wb.sheetnames:
            wb.remove(wb[name])


def _fill_step1_scores(ws: Worksheet, scores: RubricScores) -> None:
    _set_automated_score(ws, ROW_RUBRIC_ROWS["medical_school_quality"], scores.medical_school_quality)
    _set_automated_score(
        ws, ROW_RUBRIC_ROWS["medical_school_performance"], scores.medical_school_performance
    )
    _set_automated_score(ws, ROW_RUBRIC_ROWS["undergraduate_quality"], scores.undergraduate_quality)
    _set_automated_score(
        ws, ROW_RUBRIC_ROWS["undergraduate_performance"], scores.undergraduate_performance
    )
    _set_automated_score(ws, ROW_RUBRIC_ROWS["usmle_step1"], scores.usmle_step1)

    _set_summary_cell(ws, COL_SUMMARY["msq"], scores.medical_school_quality)
    _set_summary_cell(ws, COL_SUMMARY["msp"], scores.medical_school_performance)
    _set_summary_cell(ws, COL_SUMMARY["uq"], scores.undergraduate_quality)
    _set_summary_cell(ws, COL_SUMMARY["up"], scores.undergraduate_performance)
    _set_summary_cell(ws, COL_SUMMARY["usmle"], scores.usmle_step1)


def write_workbook(
    template_path: Path,
    facts: ExtractedFacts,
    scores: RubricScores,
    output_path: Path,
    template_sheet_index: int = 0,
    strip_template_sheets: bool = True,
    doc_a: AgentReview | None = None,
    doc_b: AgentReview | None = None,
) -> None:
    wb = load_workbook(template_path)
    sheetnames = wb.sheetnames
    if not sheetnames:
        raise ValueError("Template workbook has no sheets")
    original_names = list(sheetnames)
    src_name = sheetnames[min(template_sheet_index, len(sheetnames) - 1)]
    source = wb[src_name]
    title = facts.applicant_name or Path(facts.source_path).stem
    ws = _copy_sheet(wb, source, title)
    _clear_reviewer_scores(ws)
    _setup_reviewer_columns(ws)

    ws.cell(row=ROW_APPLICANT_NAME, column=COL_APPLICANT_NAME).value = facts.applicant_name
    _fill_step1_scores(ws, scores)
    _fill_agent_review(ws, COL_DOC_A, doc_a)
    _fill_agent_review(ws, COL_DOC_B, doc_b)

    if strip_template_sheets:
        _strip_original_sheets(wb, original_names)
    wb.save(output_path)


def write_multi_applicant_workbook(
    template_path: Path,
    items: list[tuple[ExtractedFacts, RubricScores]],
    output_path: Path,
    template_sheet_index: int = 0,
    strip_template_sheets: bool = True,
    agent_reviews: list[tuple[AgentReview | None, AgentReview | None]] | None = None,
) -> None:
    wb = load_workbook(template_path)
    sheetnames = wb.sheetnames
    if not sheetnames:
        raise ValueError("Template workbook has no sheets")
    original_names = list(sheetnames)
    src_name = sheetnames[min(template_sheet_index, len(sheetnames) - 1)]
    source = wb[src_name]

    for idx, (facts, scores) in enumerate(items):
        title = facts.applicant_name or Path(facts.source_path).stem
        ws = _copy_sheet(wb, source, title)
        _clear_reviewer_scores(ws)
        _setup_reviewer_columns(ws)
        ws.cell(row=ROW_APPLICANT_NAME, column=COL_APPLICANT_NAME).value = facts.applicant_name
        _fill_step1_scores(ws, scores)
        if agent_reviews and idx < len(agent_reviews):
            doc_a, doc_b = agent_reviews[idx]
            _fill_agent_review(ws, COL_DOC_A, doc_a)
            _fill_agent_review(ws, COL_DOC_B, doc_b)

    if strip_template_sheets:
        _strip_original_sheets(wb, original_names)
    wb.save(output_path)
